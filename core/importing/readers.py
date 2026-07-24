"""File-format readers that expose a common tabular sheet representation."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from core.errors import DataImportError


@dataclass(frozen=True, slots=True)
class TabularSheet:
    """Rows from one source sheet."""

    name: str | None
    rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True, slots=True)
class TabularSheetError:
    """A recoverable failure limited to one sheet inside a readable file."""

    name: str | None
    error: DataImportError


TabularReadResult = TabularSheet | TabularSheetError


class TabularReader(Protocol):
    """Structural protocol for registered readers."""

    def read(self, path: Path) -> tuple[TabularReadResult, ...]:
        """Read a file into one or more sheets."""


class CsvReader:
    """Read delimited text with UTF-8 and common Chinese encoding support."""

    _ENCODINGS = ("utf-8-sig", "gb18030", "utf-16")

    def read(self, path: Path) -> tuple[TabularSheet, ...]:
        text: str | None = None
        last_error: UnicodeError | None = None
        for encoding in self._ENCODINGS:
            try:
                text = path.read_text(encoding=encoding)
                break
            except UnicodeError as exc:
                last_error = exc
            except OSError as exc:
                raise DataImportError(
                    f"Unable to read CSV file: {path.name}",
                    code="E_IMPORT_FILE_READ",
                    detail=str(exc),
                ) from exc
        if text is None:
            raise DataImportError(
                f"Unable to decode CSV file: {path.name}",
                code="E_IMPORT_CSV_ENCODING",
                detail=str(last_error) if last_error else None,
            )

        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            rows = tuple(tuple(row) for row in csv.reader(text.splitlines(), dialect))
        except csv.Error:
            rows = tuple(tuple(row) for row in csv.reader(text.splitlines()))
        return (TabularSheet(name=None, rows=rows),)


class XlsxReader:
    """Read modern Excel workbooks through openpyxl."""

    def read(self, path: Path) -> tuple[TabularSheet, ...]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DataImportError(
                "XLSX support requires openpyxl.",
                code="E_IMPORT_DEPENDENCY",
                detail="Install dependencies from requirements.txt.",
            ) from exc

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheets = tuple(
                TabularSheet(
                    name=worksheet.title,
                    rows=tuple(tuple(row) for row in worksheet.iter_rows(values_only=True)),
                )
                for worksheet in workbook.worksheets
            )
            workbook.close()
            return sheets
        except Exception as exc:
            raise DataImportError(
                f"Unable to read Excel workbook: {path.name}",
                code="E_IMPORT_XLSX_READ",
                detail=str(exc),
            ) from exc


class XlsReader:
    """Read legacy binary Excel workbooks through xlrd."""

    def read(self, path: Path) -> tuple[TabularSheet, ...]:
        try:
            import xlrd
        except ImportError as exc:
            raise DataImportError(
                "Legacy XLS support requires xlrd.",
                code="E_IMPORT_DEPENDENCY",
                detail="Install dependencies from requirements.txt.",
            ) from exc

        try:
            workbook = xlrd.open_workbook(str(path), on_demand=True)
            sheets: list[TabularSheet] = []
            for sheet_name in workbook.sheet_names():
                worksheet = workbook.sheet_by_name(sheet_name)
                rows = tuple(
                    tuple(worksheet.cell_value(row, column) for column in range(worksheet.ncols))
                    for row in range(worksheet.nrows)
                )
                sheets.append(TabularSheet(name=sheet_name, rows=rows))
            workbook.release_resources()
            return tuple(sheets)
        except Exception as exc:
            raise DataImportError(
                f"Unable to read legacy Excel workbook: {path.name}",
                code="E_IMPORT_XLS_READ",
                detail=str(exc),
            ) from exc


class ReaderRegistry:
    """Map supported extensions to isolated reader implementations."""

    def __init__(self, origin_reader: TabularReader | None = None) -> None:
        if origin_reader is None:
            from .origin_reader import OriginProjectReader

            origin_reader = OriginProjectReader()
        self._readers: dict[str, TabularReader] = {
            ".csv": CsvReader(),
            ".xlsx": XlsxReader(),
            ".xlsm": XlsxReader(),
            ".xls": XlsReader(),
            ".opj": origin_reader,
            ".opju": origin_reader,
        }

    @property
    def supported_extensions(self) -> frozenset[str]:
        """Return recognized lowercase file extensions."""

        return frozenset(self._readers)

    def read(self, path: Path) -> tuple[TabularReadResult, ...]:
        """Dispatch a source file to its registered reader."""

        if not path.exists() or not path.is_file():
            raise DataImportError(
                f"File does not exist: {path}",
                code="E_IMPORT_FILE_NOT_FOUND",
            )
        reader = self._readers.get(path.suffix.casefold())
        if reader is None:
            raise DataImportError(
                f"Unsupported file type: {path.suffix or '(none)'}",
                code="E_IMPORT_UNSUPPORTED_FORMAT",
            )
        return reader.read(path)
