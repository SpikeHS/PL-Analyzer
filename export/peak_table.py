"""Excel and CSV exporters for raw peak table records."""

from __future__ import annotations

import csv
import os
import tempfile
from collections.abc import Sequence
from contextlib import suppress
from pathlib import Path

from core.errors import ExportError
from core.models import PeakTableRecord


class PeakTableExporter:
    """Export visible raw peak records with stable units and column order."""

    _HEADERS = (
        "Sample",
        "Material",
        "Peak",
        "Position (nm)",
        "Height (a.u.)",
        "FWHM (nm)",
        "Prominence (a.u.)",
        "Quality Flags",
    )

    def export(self, records: Sequence[PeakTableRecord], path: Path) -> None:
        """Dispatch an export based on the requested extension."""

        if not records:
            raise ExportError(
                "There are no peak results to export.",
                code="E_EXPORT_NO_RESULTS",
            )
        suffix = path.suffix.casefold()
        if suffix == ".xlsx":
            self._export_xlsx(records, path)
        elif suffix == ".csv":
            self._export_csv(records, path)
        else:
            raise ExportError(
                f"Unsupported peak table format: {path.suffix}",
                code="E_EXPORT_FORMAT",
            )

    def _export_xlsx(self, records: Sequence[PeakTableRecord], path: Path) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise ExportError(
                "Excel export requires openpyxl.",
                code="E_EXPORT_DEPENDENCY",
            ) from exc

        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = _temporary_path(path)
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Raw Peaks"
            worksheet.append(self._HEADERS)
            for record in records:
                worksheet.append(_record_values(record))
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="24527A")
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            widths = (24, 22, 9, 16, 18, 14, 20, 28)
            for column_index, width in enumerate(widths, start=1):
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=column_index).column_letter
                ].width = width
            workbook.save(temp_path)
            os.replace(temp_path, path)
        except Exception as exc:
            _remove_if_present(temp_path)
            if isinstance(exc, ExportError):
                raise
            raise ExportError(
                f"Unable to export Excel results: {path.name}",
                code="E_EXPORT_WRITE",
                detail=str(exc),
            ) from exc

    def _export_csv(self, records: Sequence[PeakTableRecord], path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = _temporary_path(path)
        try:
            with temp_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(self._HEADERS)
                for record in records:
                    writer.writerow(_record_values(record))
            os.replace(temp_path, path)
        except OSError as exc:
            _remove_if_present(temp_path)
            raise ExportError(
                f"Unable to export CSV results: {path.name}",
                code="E_EXPORT_WRITE",
                detail=str(exc),
            ) from exc


def _record_values(record: PeakTableRecord) -> tuple[object, ...]:
    return (
        record.sample_name,
        ", ".join(record.material_names),
        record.peak_number,
        record.position_nm,
        record.height_au,
        record.fwhm_nm,
        record.prominence_au,
        ", ".join(record.quality_flags),
    )


def _temporary_path(target: Path) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix=f".{target.name}.",
        suffix=".tmp",
        dir=target.parent,
    )
    os.close(descriptor)
    return Path(name)


def _remove_if_present(path: Path) -> None:
    with suppress(OSError):
        path.unlink(missing_ok=True)
