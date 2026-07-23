"""Excel and CSV export for v1.1 fitted peak metrics."""

from __future__ import annotations

import csv
import os
import tempfile
from collections.abc import Sequence
from contextlib import suppress
from pathlib import Path

from analysis.fit_session import FitTableRecord
from core.errors import ExportError


class FitTableExporter:
    """Write fitted physical metrics and fit-quality statistics."""

    _HEADERS = (
        "Sample",
        "Material",
        "Model",
        "Peak",
        "Position (nm)",
        "Height (a.u.)",
        "Area (a.u. nm)",
        "FWHM (nm)",
        "Gaussian FWHM (nm)",
        "Lorentzian FWHM (nm)",
        "Pseudo-Voigt Mixing Fraction",
        "R Squared",
        "Adjusted R Squared",
        "AIC",
        "BIC",
        "Baseline",
    )

    def export(self, records: Sequence[FitTableRecord], path: Path) -> None:
        """Export by extension."""

        if not records:
            raise ExportError(
                "There are no fit results to export.",
                code="E_EXPORT_NO_FIT_RESULTS",
            )
        if path.suffix.casefold() == ".xlsx":
            self._export_xlsx(records, path)
        elif path.suffix.casefold() == ".csv":
            self._export_csv(records, path)
        else:
            raise ExportError(
                f"Unsupported fit table format: {path.suffix}",
                code="E_EXPORT_FORMAT",
            )

    def _export_xlsx(self, records: Sequence[FitTableRecord], path: Path) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise ExportError(
                "Excel export requires openpyxl.",
                code="E_EXPORT_DEPENDENCY",
            ) from exc
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = _temporary_path(path)
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Model Fits"
            worksheet.append(self._HEADERS)
            for record in records:
                worksheet.append(_values(record))
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="5B3A8A")
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for column in worksheet.columns:
                letter = column[0].column_letter
                worksheet.column_dimensions[letter].width = min(
                    max(len(str(cell.value or "")) for cell in column) + 2,
                    30,
                )
            workbook.save(temporary)
            os.replace(temporary, path)
        except Exception as exc:
            with suppress(OSError):
                temporary.unlink(missing_ok=True)
            raise ExportError(
                f"Unable to export fit results: {path.name}",
                code="E_EXPORT_WRITE",
                detail=str(exc),
            ) from exc

    def _export_csv(self, records: Sequence[FitTableRecord], path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = _temporary_path(path)
        try:
            with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.writer(stream)
                writer.writerow(self._HEADERS)
                writer.writerows(_values(record) for record in records)
            os.replace(temporary, path)
        except OSError as exc:
            with suppress(OSError):
                temporary.unlink(missing_ok=True)
            raise ExportError(
                f"Unable to export fit results: {path.name}",
                code="E_EXPORT_WRITE",
                detail=str(exc),
            ) from exc


def _values(record: FitTableRecord) -> tuple[object, ...]:
    return (
        record.sample_name,
        record.material_name,
        record.model.value,
        record.peak_number,
        record.position_nm,
        record.height_au,
        record.area_au_nm,
        record.fwhm_nm,
        record.gaussian_fwhm_nm,
        record.lorentzian_fwhm_nm,
        record.mixing_fraction,
        record.r_squared,
        record.adjusted_r_squared,
        record.aic,
        record.bic,
        record.baseline_mode.value,
    )


def _temporary_path(target: Path) -> Path:
    descriptor, filename = tempfile.mkstemp(
        prefix=f".{target.name}.",
        suffix=".tmp",
        dir=target.parent,
    )
    os.close(descriptor)
    return Path(filename)
